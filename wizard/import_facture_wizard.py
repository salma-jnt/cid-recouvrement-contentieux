import base64
import io
import re
import unicodedata
from datetime import date, datetime

from odoo import api, fields, models, _
from odoo.exceptions import UserError

try:
    import openpyxl
    from openpyxl.utils.datetime import from_excel
except ImportError:
    openpyxl = None
    from_excel = None


class RecouvrementFactureImportWizard(models.TransientModel):
    _name = 'recouvrement.facture.import.wizard'
    _description = 'Importer des factures depuis Excel'

    file = fields.Binary(string='Fichier Excel', required=True)
    file_data = fields.Binary(string='Fichier Excel (alias)')
    file_name = fields.Char(string='Nom du fichier')
    sheet_name = fields.Char(string='Feuille (optionnel)')
    mois_upload = fields.Char(
        string="Mois d'upload",
        help="Marquage analytique au format YYYY-MM (ex: 2026-04). "
             "Renseigné par la page d'import moderne pour figer les données.",
    )
    imported_count = fields.Integer(string='Nouvelles factures', readonly=True)
    updated_count = fields.Integer(string='Factures mises à jour', readonly=True)
    locked_count = fields.Integer(string='Factures verrouillées', readonly=True)
    note = fields.Text(string='Résumé', readonly=True)

    def _normalize_header(self, header):
        if header is None:
            return ''
        value = str(header).strip().lower().replace('\r', ' ').replace('\n', ' ')
        value = value.replace('…', '...')
        value = value.replace('n°', 'numero ').replace('nº', 'numero ').replace("l'ordre", 'ordre')
        value = unicodedata.normalize('NFKD', value).encode('ascii', 'ignore').decode('ascii')
        value = re.sub(r'\s+', ' ', value)
        return value.strip(' :')

    def _get_field_mapping(self, header):
        mapping = {
            'reference facture': 'name',
            'numero facture': 'name',
            'numero de la facture': 'name',
            'facture': 'name',
            'code affaire': 'code_affaire',
            'affaire': 'code_affaire',
            'type de facture': 'facture_type',
            'type facture': 'facture_type',
            'facture type': 'facture_type',
            'client': 'client_id',
            'date de reception de lordre de facturation (1)': 'date_reception_ordre',
            'date de reception de lordre de facturation': 'date_reception_ordre',
            'chez laila': 'chez_laila',
            'chez bennis': 'chez_bennis',
            'date de la facture': 'date_facture',
            'date facture': 'date_facture',
            'la date de signature de la facture': 'date_signature',
            'date de signature': 'date_signature',
            'date signature': 'date_signature',
            'date de depot de la facture chez le client (3)': 'date_depot_client',
            'date de depot de la facture chez le client': 'date_depot_client',
            'date de depot chez le client': 'date_depot_client',
            'date de depot client': 'date_depot_client',
            'date de depot client (3)': 'date_depot_client',
            'date depot': 'date_depot_client',
            'montant facture en ttc': 'montant_ttc',
            'montant de la facture en ttc': 'montant_ttc',
            'montant facture ttc': 'montant_ttc',
            'montant ttc': 'montant_ttc',
            'montant total ttc': 'montant_ttc',
            'montant ht': 'montant_ht',
            'montant facture en ht': 'montant_ht',
            'montant de la facture ht': 'montant_ht',
            'montant de la facture en ht': 'montant_ht',
            'montant facture ht': 'montant_ht',
            'montant total ht': 'montant_ht',
            'reference justificatif': 'reference_justificatif',
            'reference du justificatif': 'reference_justificatif',
            'reference du justificatif de la facture': 'reference_justificatif',
            'reference du justificatif de la facture (dp, attachement, email client...)': 'reference_justificatif',
            'reference du justificatif de la facture (dp, attachement, email client)': 'reference_justificatif',
            'ref justificatif': 'reference_justificatif',
            'ref. justificatif': 'reference_justificatif',
            'nature': 'nature',
            'division': 'division_id',
            'pole': 'pole_id',
            'numero enreg': 'numero_enreg',
            'ice': 'ice',
            'numero marche': 'numero_marche',
            'devise': 'currency_id',
        }
        field_name = mapping.get(header)
        if field_name:
            return field_name

        if 'reference' in header and 'justificatif' in header:
            return 'reference_justificatif'
        if 'montant' in header and 'ttc' in header:
            return 'montant_ttc'
        if 'montant' in header and 'ht' in header:
            return 'montant_ht'
        if 'date' in header and 'depot' in header:
            return 'date_depot_client'

        return None

    def _extract_date_from_text(self, value, default_year=None):
        if value in [None, '']:
            return False

        text = str(value).strip()
        if not text:
            return False

        normalized = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('ascii').lower()
        normalized = normalized.replace('.', '/').replace('-', '/')
        month_map = {
            'janvier': '01', 'janv': '01',
            'fevrier': '02', 'fevr': '02', 'fev': '02',
            'mars': '03',
            'avril': '04', 'avr': '04',
            'mai': '05',
            'juin': '06',
            'juillet': '07', 'juil': '07',
            'aout': '08',
            'septembre': '09', 'sept': '09',
            'octobre': '10', 'oct': '10',
            'novembre': '11', 'nov': '11',
            'decembre': '12', 'dec': '12',
        }
        for label, month in month_map.items():
            normalized = re.sub(rf'\b{label}\b', f'/{month}/', normalized)

        normalized = re.sub(r'\s+', ' ', normalized)
        match = re.search(r'(?P<day>\d{1,2})\s*/\s*(?P<month>\d{1,2})\s*/\s*(?P<year>\d{2,4})', normalized)
        if not match:
            match = re.search(r'(?P<day>\d{1,2})\s*/\s*(?P<month>\d{1,2})(?!\s*/)', normalized)
        if not match:
            return False

        day = int(match.group('day'))
        month = int(match.group('month'))
        year_text = match.groupdict().get('year')
        year = default_year or fields.Date.today().year
        if year_text:
            year = int(year_text)
            if year < 100:
                year += 2000

        try:
            return date(year, month, day)
        except ValueError:
            return False

    def _normalize_type(self, value):
        if not value:
            return False
        value = self._normalize_header(value)
        if 'pro forma' in value or 'pro-forma' in value or 'proforma' in value:
            return 'proforma'
        if 'groupement' in value:
            return 'groupement'
        if 'avance' in value:
            return 'avance'
        if 'suivi de facturation' in value or '2eme envoi' in value or 'standard' in value:
            return 'standard'
        return 'autre'

    def _find_or_create_client(self, name):
        if not name:
            return False
        partner = self.env['res.partner'].search([('name', '=ilike', name)], limit=1)
        if partner:
            return partner
        return self.env['res.partner'].create({'name': name})

    def _find_or_create_division(self, name):
        if not name:
            return False
        division = self.env['hr.department'].search([('name', '=ilike', name)], limit=1)
        if division:
            return division
        return self.env['hr.department'].create({'name': name})

    def _find_or_create_pole(self, name):
        if not name:
            return False
        pole = self.env['hr.department'].search([('name', '=ilike', name)], limit=1)
        if pole:
            return pole
        return self.env['hr.department'].create({'name': name})

    def _convert_value(self, field_name, value):
        if value in [None, '']:
            return False

        if field_name in ['name', 'numero_enreg']:
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                return str(int(value)).zfill(3)
            return str(value).strip()

        if field_name in ['code_affaire', 'ice', 'numero_marche', 'chez_laila', 'chez_bennis', 'nature', 'reference_justificatif']:
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                return str(int(value))
            return str(value).strip()

        if field_name in ['date_reception_ordre', 'date_facture', 'date_signature', 'date_depot_client']:
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, date):
                return value
            if isinstance(value, (int, float)) and not isinstance(value, bool) and from_excel:
                try:
                    converted = from_excel(value)
                    return converted.date() if hasattr(converted, 'date') else converted
                except Exception:
                    pass
            text = str(value).strip()
            for fmt in (
                '%m/%d/%Y', '%m/%d/%y', '%d/%m/%Y', '%d/%m/%y',
                '%Y-%m-%d', '%Y-%m-%d %H:%M', '%Y-%m-%d %H:%M:%S',
                '%m/%d/%Y %H:%M', '%m/%d/%Y %H:%M:%S',
                '%d/%m/%Y %H:%M', '%d/%m/%Y %H:%M:%S'
            ):
                try:
                    return datetime.strptime(text, fmt).date()
                except ValueError:
                    continue
            extracted_date = self._extract_date_from_text(text)
            if extracted_date:
                return extracted_date
            return False

        if field_name in ['montant_ttc', 'montant_ht']:
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                return float(value)
            text = str(value).strip().replace('\xa0', '').replace(' ', '')
            if ',' in text and '.' in text:
                if text.rfind(',') > text.rfind('.'):
                    text = text.replace('.', '').replace(',', '.')
                else:
                    text = text.replace(',', '')
            elif ',' in text:
                text = text.replace(',', '.')
            try:
                return float(text)
            except (TypeError, ValueError):
                return False

        return value

    def action_import(self):
        """
        Import des factures avec respect strict des 3 cas d'upsert
        du cahier de charge (chap. 1) :

          Cas 1 — Nouvelle ligne : code_affaire inexistant
                  → CREATE avec recouvrement_status = 'normal'
                  → rattachement automatique à un dossier (groupé par
                    procedure_id + date_depot_client)

          Cas 2 — Mise à jour autorisée : facture existe ET status Normal
                  → UPDATE des champs

          Cas 3 — Verrou juridique : facture existe ET status est
                  precontentieux / contentieux / bloque_juridique / bloque_technique
                  → REJET (compte dans locked_count)
        """
        self.ensure_one()
        if openpyxl is None:
            raise UserError(_(
                'La bibliothèque Python openpyxl est requise pour importer '
                'les fichiers Excel.'
            ))

        # Source du contenu : champ legacy `file` ou nouveau `file_data`
        binary = self.file_data or self.file
        if not binary:
            raise UserError(_('Aucun fichier fourni.'))
        try:
            content = base64.b64decode(binary)
        except Exception:
            raise UserError(_('Impossible de décoder le fichier Excel.'))

        workbook = openpyxl.load_workbook(
            io.BytesIO(content), data_only=True, read_only=True,
        )
        sheet_names = [self.sheet_name] if self.sheet_name else workbook.sheetnames
        facture_obj = self.env['recouvrement.facture']
        recouvrement_obj = self.env['recouvrement.recouvrement']

        created = 0
        updated = 0
        locked = 0
        errors = []
        new_factures = self.env['recouvrement.facture']

        for sheet_name in sheet_names:
            if not sheet_name:
                continue
            if sheet_name not in workbook.sheetnames:
                errors.append(_('Feuille introuvable: %s') % sheet_name)
                continue
            worksheet = workbook[sheet_name]
            row_iterator = worksheet.iter_rows(values_only=True)

            headers = None
            header_row_index = 0
            for probe_index, probe_row in enumerate(row_iterator, start=1):
                normalized_headers = [self._normalize_header(c) for c in probe_row]
                mapped = [
                    self._get_field_mapping(h)
                    for h in normalized_headers
                    if self._get_field_mapping(h)
                ]
                if mapped and ('name' in mapped or len(mapped) >= 4):
                    headers = normalized_headers
                    header_row_index = probe_index
                    break
                if probe_index >= 20:
                    break

            if not headers:
                errors.append(_('Aucune ligne d’en-tête reconnue dans la feuille %s.') % sheet_name)
                continue

            for row_index, row in enumerate(row_iterator, start=header_row_index + 1):
                if not any([cell is not None and str(cell).strip() for cell in row]):
                    continue
                values = self._extract_row_values(headers, row, sheet_name)
                if values is None:
                    continue

                reference_value = str(values.get('name') or '').strip()
                if not reference_value or reference_value.lower() in ['vide', 'nan', 'none', '-', 'n/a']:
                    continue
                values['name'] = reference_value

                if not values.get('client_id'):
                    errors.append(_(
                        'Ligne %(row)s feuille %(sheet)s ignorée : client manquant.',
                        row=row_index, sheet=sheet_name,
                    ))
                    continue

                # Marquage analytique cahier
                if self.mois_upload:
                    values['mois_upload'] = self.mois_upload

                values['source_sheet'] = sheet_name

                # ============================================================
                # LOGIQUE D'UPSERT — 3 cas du cahier
                # ============================================================
                code_affaire = (values.get('code_affaire') or '').strip()
                facture_existante = False

                if code_affaire:
                    # Recherche par code_affaire (clé unique métier per cahier)
                    facture_existante = facture_obj.search(
                        [('code_affaire', '=', code_affaire)], limit=1,
                    )
                else:
                    # Fallback : recherche par numéro de facture
                    facture_existante = facture_obj.search(
                        [('name', '=', reference_value)], limit=1,
                    )

                try:
                    if facture_existante:
                        # ----- Cas 2 ou Cas 3 -----
                        if facture_existante.recouvrement_status in (
                            'precontentieux', 'contentieux',
                            'bloque_juridique', 'bloque_technique',
                        ):
                            # CAS 3 : verrou juridique → REJET
                            locked += 1
                            errors.append(_(
                                'Verrou juridique : %(name)s (statut %(status)s) — non mise à jour.',
                                name=facture_existante.name,
                                status=facture_existante.recouvrement_status,
                            ))
                            continue

                        # CAS 2 : statut Normal → UPDATE autorisée
                        update_vals = {
                            k: v for k, v in values.items()
                            if v is not None and v != '' and v is not False
                        }
                        if update_vals:
                            facture_existante.with_context(
                                recouvrement_import_unlock=True,
                            ).write(update_vals)
                            updated += 1
                            
                        # Si la facture existante n'est rattachée à aucun dossier, 
                        # on l'ajoute au pipeline pour la grouper dans un convoi
                        if not facture_existante.recouvrement_id:
                            new_factures |= facture_existante
                    else:
                        # ----- Cas 1 : NOUVELLE FACTURE -----
                        defaults = {
                            'state': 'imported',
                            'recouvrement_status': 'normal',
                            'statut_interface': 'vert',
                            'currency_id': self.env.company.currency_id.id,
                        }
                        defaults.update({
                            k: v for k, v in values.items()
                            if v is not None and v != '' and v is not False
                        })
                        new_facture = facture_obj.create(defaults)
                        new_factures |= new_facture
                        created += 1
                except Exception as exc:
                    errors.append(_(
                        'Ligne %(row)s feuille %(sheet)s : %(msg)s',
                        row=row_index, sheet=sheet_name, msg=str(exc),
                    ))
                    continue

        # ============================================================
        # Rattachement automatique à un dossier (logique du convoi)
        # ============================================================
        for facture in new_factures:
            if facture.recouvrement_id:
                continue
            try:
                dossier = recouvrement_obj.get_or_create_for_facture(facture)
                if dossier:
                    facture.with_context(
                        recouvrement_import_unlock=True,
                    ).write({'recouvrement_id': dossier.id})
            except Exception as exc:
                errors.append(_(
                    "Rattachement dossier facture %(name)s : %(msg)s",
                    name=facture.name, msg=str(exc),
                ))

        # ============================================================
        # Résumé & retour
        # ============================================================
        self.imported_count = created
        self.updated_count = updated
        self.locked_count = locked

        note_lines = [_('Import terminé.')]
        note_lines.append(_('  • Nouvelles factures : %s') % created)
        note_lines.append(_('  • Factures mises à jour : %s') % updated)
        if locked:
            note_lines.append(_('  • Factures verrouillées (cas 3) : %s') % locked)
        if errors:
            note_lines.append('')
            note_lines.append(_('Avertissements / erreurs :'))
            note_lines.extend(['  - ' + e for e in errors[:50]])
            if len(errors) > 50:
                note_lines.append(_('  ... et %s autres') % (len(errors) - 50))
        self.note = '\n'.join(note_lines)

        # Retour : si appelé depuis la page OWL, on retourne le résumé brut.
        # Si appelé depuis le wizard standard, on rouvre le wizard avec note.
        if self.env.context.get('from_owl_page'):
            return {
                'created': created,
                'updated': updated,
                'locked': locked,
                'total': created + updated + locked,
                'errors': errors,
            }
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'recouvrement.facture.import.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }

    # ==================================================================
    # Helper : extraction d'une ligne en dict de valeurs Odoo
    # ==================================================================
    def _extract_row_values(self, headers, row, sheet_name):
        """Convertit une ligne (tuple) en dict prêt pour create/write."""
        values = {}
        for header, cell in zip(headers, row):
            field_name = self._get_field_mapping(header)
            if not field_name:
                continue

            # Champs supprimés en v2 : on les ignore proprement
            if field_name in ('chez_laila', 'chez_bennis'):
                continue

            if field_name == 'client_id':
                client = self._find_or_create_client(cell)
                if client:
                    values['client_id'] = client.id
                continue
            if field_name == 'division_id':
                division = self._find_or_create_division(cell)
                if division:
                    values['division_id'] = division.id
                continue
            if field_name == 'pole_id':
                pole = self._find_or_create_pole(cell)
                if pole:
                    values['pole_id'] = pole.id
                continue
            if field_name == 'currency_id' and cell:
                currency = self.env['res.currency'].search(
                    [('name', '=ilike', str(cell).strip())], limit=1,
                )
                if currency:
                    values['currency_id'] = currency.id
                continue
            if field_name == 'facture_type':
                type_value = self._normalize_type(cell)
                if type_value:
                    values['facture_type'] = type_value
                continue

            converted_value = self._convert_value(field_name, cell)
            if field_name == 'date_depot_client':
                if converted_value:
                    values['date_depot_client'] = converted_value
                else:
                    # tentative forcée depuis texte
                    extracted = self._extract_date_from_text(cell)
                    if extracted:
                        values['date_depot_client'] = extracted
                    else:
                        values['depot_comment'] = str(cell).strip()
                continue
            values[field_name] = converted_value

        # Type de facture déduit de la feuille si pas explicite
        type_from_sheet = self._normalize_type(sheet_name)
        if type_from_sheet and not values.get('facture_type'):
            values['facture_type'] = type_from_sheet

        # Nettoyage code_affaire vide
        if 'code_affaire' in values and not values.get('code_affaire'):
            values.pop('code_affaire')

        return values
