import base64
import io
from datetime import date, datetime

from odoo import _, fields, models
from odoo.exceptions import UserError

try:
    import openpyxl
    from openpyxl.utils.datetime import from_excel
except ImportError:
    openpyxl = None
    from_excel = None


class RecouvrementEncaissementImportWizard(models.TransientModel):
    _name = 'recouvrement.encaissement.import.wizard'
    _description = 'Importer des encaissements depuis Excel'

    file = fields.Binary(string='Fichier Excel', required=True)
    file_name = fields.Char(string='Nom du fichier')
    imported_count = fields.Integer(string='Nouveaux encaissements', readonly=True)
    updated_count = fields.Integer(string='Encaissements mis a jour', readonly=True)
    note = fields.Text(string='Resume', readonly=True)

    def _normalize_header(self, header):
        if header is None:
            return ''
        value = str(header).strip().lower().replace('\r', ' ').replace('\n', ' ')
        value = value.replace('n°', 'numero ').replace('nº', 'numero ')
        value = value.replace('…', '...')
        return ' '.join(value.split()).strip(' :')

    def _get_field_mapping(self, header):
        mapping = {
            'reference': 'name',
            'reference encaissement': 'name',
            'numero encaissement': 'name',
            'facture': 'facture_ref',
            'numero facture': 'facture_ref',
            'code affaire': 'code_affaire',
            'montant': 'montant',
            'montant encaisse': 'montant',
            'penalite': 'penalite',
            'date operation': 'date_operation',
            'date d operation': 'date_operation',
            'date operation encaissement': 'date_operation',
            'banque': 'banque',
            'observation': 'observation',
            'commentaire': 'observation',
            'mode paiement': 'mode_paiement',
            'mode de paiement': 'mode_paiement',
        }
        field_name = mapping.get(header)
        if field_name:
            return field_name
        if 'date' in header and 'operation' in header:
            return 'date_operation'
        return None

    def _parse_date(self, value):
        if value in [None, '']:
            return False
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)) and not isinstance(value, bool) and from_excel:
            try:
                converted = from_excel(value)
                return converted.date() if hasattr(converted, 'date') else converted
            except Exception:
                pass
        text = str(value).strip()
        for fmt in (
            '%Y-%m-%d', '%Y-%m-%d %H:%M', '%Y-%m-%d %H:%M:%S',
            '%d/%m/%Y', '%d/%m/%Y %H:%M', '%d/%m/%Y %H:%M:%S',
            '%m/%d/%Y', '%m/%d/%Y %H:%M', '%m/%d/%Y %H:%M:%S',
        ):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return False

    def _parse_float(self, value):
        if value in [None, '']:
            return False
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return float(value)
        text = str(value).strip().replace('\xa0', '').replace(' ', '')
        if ',' in text and '.' in text:
            if text.rfind(',') > text.rfind('.'):
                text = text.replace('.', '').replace(',', '.')
            else:
                text = text.replace(',', '')
        elif ',' in text:
            text = text.replace(',', '.')
        try:
            return float(text)
        except (TypeError, ValueError):
            return False

    def _parse_mode(self, value):
        if not value:
            return 'virement'
        normalized = str(value).strip().lower()
        if 'espece' in normalized:
            return 'espece'
        if 'cheque' in normalized:
            return 'cheque'
        if 'vir' in normalized:
            return 'virement'
        return 'autre'

    def action_import(self):
        self.ensure_one()
        if openpyxl is None:
            raise UserError(_('La bibliotheque Python openpyxl est requise pour importer les fichiers Excel.'))

        workbook = openpyxl.load_workbook(io.BytesIO(base64.b64decode(self.file)), data_only=True, read_only=True)
        encaissement_obj = self.env['recouvrement.encaissement']
        facture_obj = self.env['recouvrement.facture']
        created = 0
        updated = 0
        errors = []

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            row_iterator = worksheet.iter_rows(values_only=True)
            headers = None
            for row in row_iterator:
                normalized = [self._normalize_header(cell) for cell in row]
                if any(self._get_field_mapping(h) for h in normalized):
                    headers = normalized
                    break
            if not headers:
                continue

            for row_index, row in enumerate(row_iterator, start=2):
                if not any(cell not in [None, ''] for cell in row):
                    continue
                values = {}
                facture_ref = False
                code_affaire = False

                for header, cell in zip(headers, row):
                    field_name = self._get_field_mapping(header)
                    if not field_name:
                        continue
                    if field_name == 'facture_ref':
                        facture_ref = str(cell).strip() if cell not in [None, ''] else False
                        continue
                    if field_name == 'code_affaire':
                        code_affaire = str(cell).strip() if cell not in [None, ''] else False
                        continue
                    if field_name == 'date_operation':
                        values['date_operation'] = self._parse_date(cell)
                        continue
                    if field_name in ('montant', 'penalite'):
                        values[field_name] = self._parse_float(cell)
                        continue
                    if field_name == 'mode_paiement':
                        values['mode_paiement'] = self._parse_mode(cell)
                        continue
                    values[field_name] = str(cell).strip() if cell not in [None, ''] else False

                if not values.get('date_operation') or values.get('montant') in [False, None]:
                    errors.append(_('Ligne %s ignoree: date operation ou montant manquant.') % row_index)
                    continue

                facture = False
                if facture_ref:
                    domain = [('name', '=', facture_ref)]
                    if code_affaire:
                        domain.append(('code_affaire', '=', code_affaire))
                    facture = facture_obj.search(domain, limit=1)

                if facture:
                    values['facture_id'] = facture.id
                    values['recouvrement_id'] = facture.recouvrement_id.id if facture.recouvrement_id else False

                search_domain = [('name', '=', values.get('name'))] if values.get('name') else []
                if search_domain and values.get('date_operation'):
                    search_domain.append(('date_operation', '=', values['date_operation']))
                existing = encaissement_obj.search(search_domain, limit=1) if search_domain else False

                vals_to_write = {k: v for k, v in values.items() if v is not None and v != '' and v is not False}
                try:
                    if existing:
                        existing.write(vals_to_write)
                        updated += 1
                    else:
                        if not vals_to_write.get('name'):
                            vals_to_write['name'] = 'Nouveau'
                        encaissement_obj.create(vals_to_write)
                        created += 1
                except Exception as exc:
                    errors.append(_('Ligne %s ignoree: %s') % (row_index, str(exc)))

        self.imported_count = created
        self.updated_count = updated
        summary = [_('Import termine.'), _('Nouveaux encaissements: %s') % created, _('Encaissements mis a jour: %s') % updated]
        summary.extend(errors)
        self.note = '\n'.join(summary)

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'recouvrement.encaissement.import.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }
